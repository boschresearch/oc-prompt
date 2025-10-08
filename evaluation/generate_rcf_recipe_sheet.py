#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Copyright (c) 2025 Robert Bosch GmbH
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published
# by the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

import json
import logging
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


def main():
    # Determine project root (two levels up from this script)
    project_root = Path(__file__).resolve().parent.parent

    # Load selection configuration
    config_path = project_root / "data" / "recipe" / "config" / "rcf_recipe_selection.json"
    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)

    recipe_ids = config.get("recipe_ids", [])
    techniques = config.get("prompting_techniques", [])
    llm_names = config.get("llm_names", [])

    # Prepare output directory and logging
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    eval_dir = project_root / "data" / "recipe" / "evaluation" / "RCF" / timestamp
    eval_dir.mkdir(parents=True, exist_ok=True)
    log_file = eval_dir / "generate_recipe_excel.log"
    logging.basicConfig(
        filename=str(log_file),
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s"
    )

    records = []
    base_dir = project_root / "data" / "recipe" / "target"

    # Iterate over recipes, techniques, and LLMs
    for recipe_id in recipe_ids:
        for technique in techniques:
            for llm in llm_names:
                recipe_base = base_dir / technique / llm
                pattern = f"{recipe_id}_*"
                dirs = list(recipe_base.glob(pattern))
                if not dirs:
                    logging.warning(f"No directory for {technique}_{recipe_id} under {llm}")
                    continue
                subdir = dirs[0]
                json_file = subdir / f"{subdir.name}.json"
                if not json_file.exists():
                    logging.warning(f"Missing JSON file {json_file}")
                    continue
                try:
                    with open(json_file, "r", encoding="utf-8") as rf:
                        data = json.load(rf)
                    # Extract dish name and recipe text
                    dish_name = data.get("english_target_dish_name", "")
                    recipe_body = data.get("english_target_recipe", "")
                    combined_text = f"{dish_name}\n{recipe_body}"

                    records.append({
                        "Recipe ID": f"{technique}_{recipe_id}",
                        "Recipe Text": combined_text,
                        "Expert Score": "",
                        "Expert Notes": ""
                    })
                    logging.info(f"Loaded {technique}_{recipe_id} from {json_file}")
                except Exception as e:
                    logging.error(f"Error reading {json_file}: {e}")

    # Build DataFrame
    df = pd.DataFrame(records, columns=["Recipe ID", "Recipe Text", "Expert Score", "Expert Notes"])

    # Save to Excel file
    excel_path = eval_dir / "recipes_evaluation.xlsx"
    df.to_excel(excel_path, index=False)

    # Apply styling and data validation
    wb = load_workbook(excel_path)
    ws = wb.active
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Bold header and borders
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            if cell.row == 1:
                cell.font = Font(bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 150
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30

    # Wrap text for Recipe Text
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Create dropdown for Expert Score (0 to 10, 0.5 steps)
    score_values = [str(i * 0.5) for i in range(0, 21)]
    dv = DataValidation(type="list", formula1='"' + ','.join(score_values) + '"', allow_blank=True)
    dv.error = 'Select a value from the list.'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please select a score between 0 and 10 in 0.5 increments.'
    dv.promptTitle = 'Expert Score'
    ws.add_data_validation(dv)
    max_row = ws.max_row
    # Use dv.add instead of dv.ranges.append to support current openpyxl versions
    dv.add(f"C2:C{max_row}")

    wb.save(excel_path)
    logging.info(f"Saved Excel file to {excel_path}")

if __name__ == "__main__":
    main()
